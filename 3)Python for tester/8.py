#faking data in different language
from faker import Faker
from openpyxl import Workbook
fake=Faker(['hi-IN','en-US'])
x=Workbook()
a=x.active
for i in range(1,11):
    a.cell(row=i,column=1,value=fake.name())
    a.cell(row=i,column=2,value=fake.email())
    a.cell(row=i,column=3,value=fake.address())
    a.cell(row=i,column=4,value=fake.phone_number())
x.save('fake_data_diff_language.xlsx')