#writing the data into excel sheet
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 'Name' #writes the data'Name' in the a1 cell of excel
ws['B1'] = 'Age'  #writes the data 'Age' in the b1 cell of excel
wb.save("sample.xlsx")