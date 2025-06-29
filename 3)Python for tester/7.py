#fake data for multiple cooumns
from faker import Faker
from openpyxl import Workbook
x=Workbook()
a=x.active
f=Faker()
for i in range(1,11):
        a.cell(row=i,column=1,value=f.name())
        a.cell(row=i,column=2,value=f.email())
        a.cell(row=i,column=3,value=f.address())
x.save("fake_data_multi_coloumns.xlsx")